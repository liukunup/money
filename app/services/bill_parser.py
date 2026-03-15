import csv
import io
from datetime import datetime
from decimal import Decimal
from typing import List, Optional
from pathlib import Path

from app.models.category import Category

# 类别关键词映射
CATEGORY_KEYWORDS = {
    "餐饮": ["午餐", "晚餐", "早餐", "外卖", "餐厅", "麦当劳", "肯德基", "星巴克", "瑞幸", "奶茶", "小吃", "烧烤", "火锅", "冒菜", "黄焖鸡", "沙县", "兰州拉面", "便利店", "全家", "罗森", "7-11"],
    "交通": ["地铁", "公交", "出租车", "滴滴", "加油", "停车", "高速", "打车", "网约车", "共享单车", "火车", "高铁", "飞机", "长途"],
    "购物": ["淘宝", "京东", "拼多多", "天猫", "超市", "苏宁", "国美", "唯品会", "蘑菇街", "网易严选", "小米", "苹果", "华为"],
    "娱乐": ["电影", "游戏", "KTV", "音乐", "视频", "爱奇艺", "优酷", "腾讯视频", "B站", "网易云", "QQ音乐", "健身", "游泳", "羽毛球", "篮球", "足球"],
    "住房": ["房租", "水电", "燃气", "物业", "宽带", "话费", "中介", "维修", "保洁"],
    "医疗": ["药店", "医院", "门诊", "体检", "保险", "医保", "牙科", "眼科"],
    "教育": ["学费", "培训", "课程", "书籍", "文具", "教育", "辅导", "留学"],
    "工资": ["工资", "薪资", "月薪", "奖金", "补贴"],
    "投资": ["理财", "基金", "股票", "债券", "收益", "分红", "利息"],
    "兼职": ["兼职", "外快", "项目", "佣金"],
}


class ParsedTransaction:
    """解析后的交易记录"""
    def __init__(self, amount: Decimal, trans_type: str, date: datetime, note: str, category_id: Optional[int] = None):
        self.amount = amount
        self.type = trans_type  # 'income' or 'expense'
        self.date = date
        self.note = note
        self.category_id = category_id


class BillParser:
    """账单解析器 - 支持Alipay CSV, WeChat CSV, Excel"""
    
    def __init__(self, db_session):
        self.db = db_session
        self._load_categories()
    
    def _load_categories(self):
        """加载分类用于关键词匹配"""
        categories = self.db.query(Category).filter(Category.is_deleted == False).all()
        self.categories_map = {cat.name: cat.id for cat in categories}
        self.categories_type_map = {cat.name: cat.type for cat in categories}
    
    def parse(self, file_content: bytes, filename: str, file_type: str = None) -> List[ParsedTransaction]:
        """解析文件内容"""
        ext = Path(filename).suffix.lower()
        
        # 自动检测文件类型
        if file_type:
            pass
        elif ext in ['.xlsx', '.xls']:
            file_type = 'excel'
        else:
            # 尝试通过内容检测CSV类型
            file_type = self._detect_csv_type(file_content)
        
        if file_type == 'alipay_csv':
            return self._parse_alipay_csv(file_content)
        elif file_type == 'wechat_csv':
            return self._parse_wechat_csv(file_content)
        elif file_type == 'excel':
            return self._parse_excel(file_content)
        else:
            raise ValueError(f"不支持的文件格式: {file_type}")
    
    def _detect_csv_type(self, content: bytes) -> str:
        """通过内容检测CSV类型"""
        text = content.decode('utf-8-sig', errors='ignore')
        lines = text.split('\n')[:5]
        
        # Alipay CSV特征
        alipay_markers = ['支付宝', '交易对方', '商品说明', '商家实收']
        # WeChat CSV特征  
        wechat_markers = ['微信支付', '交易对方', '商品', '当前状态']
        
        text_content = '\n'.join(lines)
        
        if any(m in text_content for m in alipay_markers):
            return 'alipay_csv'
        elif any(m in text_content for m in wechat_markers):
            return 'wechat_csv'
        
        return 'unknown'
    
    def _parse_alipay_csv(self, content: bytes) -> List[ParsedTransaction]:
        """解析支付宝CSV"""
        transactions = []
        text = content.decode('utf-8-sig', errors='ignore')
        
        reader = csv.DictReader(io.StringIO(text))
        
        for row in reader:
            try:
                # 支付宝字段映射
                amount_str = row.get('金额', row.get('金额(元)', '')).strip()
                if not amount_str:
                    continue
                
                amount = abs(Decimal(amount_str))
                
                # 判断收支类型
                amount_field = row.get('金额', row.get('金额(元)', ''))
                is_income = '收入' in row.get('业务类型', '') or Decimal(amount_field) > 0
                
                # 日期解析
                date_str = row.get('创建时间', row.get('完成时间', '')).strip()
                if date_str:
                    date = self._parse_date(date_str)
                else:
                    continue
                
                # 备注
                note = row.get('商品说明', row.get('备注', '')).strip()
                
                # 自动分类
                category_id = self._auto_categorize(note, 'income' if is_income else 'expense')
                
                transactions.append(ParsedTransaction(
                    amount=amount,
                    trans_type='income' if is_income else 'expense',
                    date=date,
                    note=note,
                    category_id=category_id
                ))
            except Exception:
                continue
        
        return transactions
    
    def _parse_wechat_csv(self, content: bytes) -> List[ParsedTransaction]:
        """解析微信CSV"""
        transactions = []
        text = content.decode('utf-8-sig', errors='ignore')
        
        reader = csv.DictReader(io.StringIO(text))
        
        for row in reader:
            try:
                # 微信字段映射
                amount_str = row.get('金额(元)', row.get('金额', '')).strip()
                if not amount_str:
                    continue
                
                amount = abs(Decimal(amount_str))
                
                # 判断收支类型
                is_income = row.get('交易类型', '') in ['收入', '转账', '红包'] or \
                           '收款' in row.get('交易状态', '')
                
                # 日期解析
                date_str = row.get('交易时间', '').strip()
                if date_str:
                    date = self._parse_date(date_str)
                else:
                    continue
                
                # 备注
                note = row.get('商品', row.get('备注', row.get('交易对方', ''))).strip()
                
                # 自动分类
                category_id = self._auto_categorize(note, 'income' if is_income else 'expense')
                
                transactions.append(ParsedTransaction(
                    amount=amount,
                    trans_type='income' if is_income else 'expense',
                    date=date,
                    note=note,
                    category_id=category_id
                ))
            except Exception:
                continue
        
        return transactions
    
    def _parse_excel(self, content: bytes) -> List[ParsedTransaction]:
        """解析Excel文件"""
        transactions = []
        
        # 使用openpyxl直接读取
        from openpyxl import load_workbook
        
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        
        # 获取表头
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # 寻找金额列
        amount_col = None
        for col in ['金额', '金额(元)', 'amount', 'Amount']:
            if col in headers:
                amount_col = headers.index(col)
                break
        
        # 寻找日期列
        date_col = None
        for col in ['日期', '交易日期', '时间', 'date', 'Date', '创建时间']:
            if col in headers:
                date_col = headers.index(col)
                break
        
        # 寻找备注列
        note_col = None
        for col in ['备注', '说明', '商品', '描述', 'note', 'Note', '商品说明']:
            if col in headers:
                note_col = headers.index(col)
                break
        
        if amount_col is None or date_col is None:
            raise ValueError("Excel文件缺少必要列(金额、日期)")
        
        # 从第二行开始读取数据
        for row_idx in range(2, ws.max_row + 1):
            try:
                amount_cell = ws.cell(row=row_idx, column=amount_col + 1)
                amount_str = str(amount_cell.value).strip() if amount_cell.value else ''
                
                if not amount_str or amount_str == 'None':
                    continue
                
                amount = abs(Decimal(amount_str))
                
                # 判断收支类型（通过金额正负）
                is_income = Decimal(str(amount_cell.value)) > 0
                
                # 日期解析
                date_cell = ws.cell(row=row_idx, column=date_col + 1)
                date = self._parse_date(str(date_cell.value))
                
                # 备注
                note = ''
                if note_col is not None:
                    note_cell = ws.cell(row=row_idx, column=note_col + 1)
                    note = str(note_cell.value) if note_cell.value else ''
                    if note == 'None':
                        note = ''
                
                # 自动分类
                category_id = self._auto_categorize(note, 'income' if is_income else 'expense')
                
                transactions.append(ParsedTransaction(
                    amount=amount,
                    trans_type='income' if is_income else 'expense',
                    date=date,
                    note=note,
                    category_id=category_id
                ))
            except Exception:
                continue
        
        return transactions
    
    def _parse_date(self, date_str: str) -> datetime:
        """解析日期字符串"""
        if not date_str or date_str == 'None':
            return datetime.now()
        
        # 尝试多种格式
        formats = [
            '%Y-%m-%d %H:%M:%S',
            '%Y/%m/%d %H:%M:%S',
            '%Y-%m-%d',
            '%Y/%m/%d',
            '%Y年%m月%d日 %H:%M:%S',
            '%Y年%m月%d日',
            '%m/%d/%Y %H:%M:%S',
            '%m/%d/%Y',
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue
        
        # 默认返回今天
        return datetime.now()
    
    def _auto_categorize(self, note: str, trans_type: str) -> Optional[int]:
        """基于关键词自动分类"""
        if not note:
            return None
        
        # 首先尝试精确匹配分类名
        for cat_name, cat_id in self.categories_map.items():
            if cat_name in note:
                if self.categories_type_map.get(cat_name) == trans_type:
                    return cat_id
        
        # 然后尝试关键词匹配
        keywords = CATEGORY_KEYWORDS
        for cat_name, kw_list in keywords.items():
            if self.categories_type_map.get(cat_name) != trans_type:
                continue
            
            for kw in kw_list:
                if kw in note:
                    return self.categories_map.get(cat_name)
        
        return None
